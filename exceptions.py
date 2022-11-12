from openpyxl import load_workbook


# напишим функцию которая проверяет заполненность полей
def check_empty_data_table(path):
    book = load_workbook(filename=path)
    shem_table = book['S2T таблицы']
    values = [
        shem_table['E' + str(3)].value[0:9],
        shem_table['E' + str(4)].value[0:4],
        shem_table['F' + str(3)].value[0:4],
        shem_table['G' + str(3)].value[0:4],
    ]
    if 'RDV_DICT.' == values[0]:
        z1 = True
    else:
        z1 = False
    if 'RDV.' == values[1]:
        z2 = True
    else:
        z2 = False
    if 'IDL.' == values[2]:
        z3 = True
    else:
        z3 = False
    if 'BDM.' == values[3]:
        z4 = True
    else:
        z4 = False
    if (shem_table['D' + str(3)].value == None) or (shem_table['D' + str(4)].value == None):
        z5 = False
    else:
        z5 = True
    res = [z1, z2, z3, z4, z5]
    return res


def check_empty_data_param(path):
    book = load_workbook(filename=path)
    shem_param = book['S2T поля']
    i = 3
    item = i
    count_PK1 = 0
    count_PK2 = 0
    z0 = True
    zC = True
    zD = True
    zE = True
    zI = True
    while (shem_param['E' + str(i)].value != None):
        if (shem_param['E' + str(i)].value == None) and (shem_param['E' + str(i + 1)].value != None):
            z0 = False
        i = i + 1
    for j in range(3, i):
        if shem_param['C' + str(j)].value == None:
            zC = False
    for j in range(3, i):
        if shem_param['D' + str(j)].value == None:
            zD = False
    for j in range(3, i):
        if shem_param['E' + str(j)].value == None:
            zE = False
    for j in range(3, i):
        if shem_param['I' + str(j)].value == None:
            zI = False
    while (shem_param['C' + str(item)].value == 1):
        if ((shem_param['J' + str(item)].value == 'PK') or (shem_param['J' + str(item)].value == 'PK (History)')):
            count_PK1 = count_PK1 + 1
        item = item + 1
    while (shem_param['C' + str(item)].value == 2):
        if ((shem_param['J' + str(item)].value == 'PK') or (shem_param['J' + str(item)].value == 'PK (History)')):
            count_PK2 = count_PK2 + 1
        item = item + 1
    return z0, zC, zD, zE, zI, count_PK1 >= 1, count_PK2 >= 1


class Data_Error(Exception):
    pass


def get_exception(path):
    for i in range(5):
        if check_empty_data_table(path)[i] == False:
            raise Data_Error("Убедитесь что в листе 'S2T таблицы' сущности названы корректно ")
    for j in range(7):
        if check_empty_data_param(path)[j] == False:
            raise Data_Error("Убедитесь что в листе 'S2T поля' все атрибуты, типы данных, истчники и РК заполнены")
