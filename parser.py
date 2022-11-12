import os.path

from openpyxl import Workbook
from export_file import array_sripts_rdv, get_array_objects, array_sripts_rdv_dict, array_sripts_IDL, array_sripts_BDM
from exceptions import get_exception


def parse(PATH: str, folder: str):
    get_exception(PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "Проверки RDV"
    ws.cell(1, 1, "Название")
    ws.cell(1, 2, "Действие")
    ws.cell(1, 3, "Ожидаемый результат")
    name_cheks = ['Проверка доступности таблицы', 'Проверка наличия данных в справочниках', 'Проверка атрибутивного состава',
                  'Проверка логических дублей', 'Проверка отсутствия старых объектов в базе', 'Проверка общего кол-ва строк эталона и мапа',
                  'Проверка истории (есть ли невалидные интервалы)', 'Проверка нахождения дат в корректном интервале', 'Проверка наличия аксессора типа "Вью"',
                  'Проверка наличия аксессора типа "Inline-функция"', 'Запуск аксессора v_sn_vld',
                  'Запуск аксессора v_sv_vld', 'Запуск аксессора v_iv_vld']
    rezults_cheks = ['Запрос возвращает записи', 'Запрос возвращает значение true', 'Запрос не возвращает записи', 'Запрос не возвращает записи',
                     'Запрос возвращает значение true', 'Запрос возвращает значение 0', 'Запрос не возвращает записи', 'Запрос не возвращает записи',
                     'Запрос возвращает значение true', 'Запрос возвращает значение true', 'Запрос не возвращает записи', 'Запрос не возвращает записи',
                     'Запрос не возвращает записи']
    name_checks_IDL = ['Проверка доступности таблицы', 'Проверка наличия данных в справочниках', 'Проверка атрибутивного состава',
                       'Заполнение src_cd', 'Проверка пересечений и разрывов в истории', 'Проверка отсутствия старых объектов в базе']
    rezults_cheks_IDL = ['Запрос возвращает записи', 'Запрос возвращает записи', 'Запрос не возвращает записи',
                         'Запрос не возвращает записи', 'Запрос не возвращает записи', 'Запрос возвращает значение true']
    name_checks_BDM = ['Проверка доступности таблицы', 'Проверка наличия данных в справочниках', 'Проверка наличия аксессоров типа "Вью"',
                       'Проверка наличия аксессоров типа "Inline-функция"', 'Проверка запуска аксессора типа "Вью" (v_sn_all_dict)',
                       'Проверка запуска аксессора типа "Вью" (v_r_dict)', 'Проверка запуска аксессора типа "Вью" (dict)',
                       'Проверка запуска аксессора типа "Inline-функция" (v_sv_all_dict)', '	Проверка запуска аксессора типа "Inline-функция" (v_sv_hash_dict)',
                       'Проверка запуска аксессора типа "Inline-функция" (v_iv_all_dict)']
    rezults_cheks_BDM = ['Запрос возвращает записи', 'Запрос возвращает записи', 'Запрос возвращает значение true',
                         'Запрос возвращает значение true', 'Запрос не возвращает записи', 'Запрос не возвращает записи',
                         'Запрос не возвращает записи', 'Запрос не возвращает записи', 'Запрос не возвращает записи', 'Запрос не возвращает записи']
    ws.cell(2, 1, get_array_objects(PATH)[0])
    ws.cell(15, 1, get_array_objects(PATH)[1])
    ws.cell(23, 1, get_array_objects(PATH)[2])
    ws.cell(29, 1, get_array_objects(PATH)[3])
    for i in range(2, 12):
        ws.cell(i, 2, '/*' + name_cheks[i - 2] + '*/ <pre class="ql-syntax" spellcheck="false">' + array_sripts_rdv(PATH)[i - 2] + "</pre>")
    for i in range(0, 3):
        ws.cell(i + 12, 2, '/*' + name_cheks[i + 10] + '*/ <pre class="ql-syntax" spellcheck="false">' + array_sripts_rdv(PATH)[10][i] + "</pre>")
    for i in range(2, 15):
        ws.cell(i, 3, rezults_cheks[i - 2])
    for i in range(0, 8):
        ws.cell(i + 15, 2, '/*' + name_cheks[i] + '*/ <pre class="ql-syntax" spellcheck="false">' + array_sripts_rdv_dict(PATH)[i] + "</pre>")
        ws.cell(i + 15, 3, rezults_cheks[i])
    for i in range(0, 6):
        ws.cell(i + 23, 2, '/*' + name_checks_IDL[i] + '*/ <pre class="ql-syntax" spellcheck="false">' + array_sripts_IDL(PATH)[i] + "</pre>")
        ws.cell(i + 23, 3, rezults_cheks_IDL[i])
    for i in range(0, 10):
        ws.cell(i + 29, 2, '/*' + name_checks_BDM[i] + '*/ <pre class="ql-syntax" spellcheck="false">' + array_sripts_BDM(PATH)[i] + "</pre>")
        ws.cell(i + 29, 3, rezults_cheks_BDM[i])
    print(array_sripts_rdv(PATH)[0])

    new_filename = 'cheks103.xlsx'
    wb.save(os.path.join(folder, new_filename))

    return new_filename
