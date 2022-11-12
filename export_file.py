from openpyxl import load_workbook, Workbook


# Проверки для слоя RDV

# 1 Проверка доступности таблицы
def check_true_schems_rdv(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        "'" + shem['E' + str(3)].value[13:29] + "'",
        "'" + shem['E' + str(4)].value[4:24] + "'",
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    result = "select table_name from information_schema.tables where 1=1 \n" \
             "and table_schema = 'rdv' \n" \
             "and table_name = " + values[1] + "\n" \
                                               "and table_type ='VIEW'"
    return result


# 2  Проверка наличия данных в справочниках
def check_data_in_directories_rdv(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        "rdv.map_" + shem['E' + str(4)].value[8:24] + "",
        "idl." + shem['F' + str(3)].value[4:25] + "",
        "bdm." + shem['G' + str(3)].value[4:25] + "",
    ]
    results_rdv = "select count(*)>1 from " + values[1] + " where src_cd = 'ACPD' and src_system='" + shem[
        'D' + str(4)].value + "';"
    results_rdv_dict = "select * from " + values[0] + " where src_cd = 'ACPD' and src_system=" + shem[
        'D' + str(3)].value + ";"
    results_idl = "select * from " + values[2] + " where src_cd = 'ACPD' and src_system=" + shem[
        'D' + str(3)].value + ";"
    results_bdm = "select * from " + values[3] + " where src_cd = 'ACPD' and src_system=" + shem[
        'D' + str(3)].value + ";"
    return results_rdv


# 3 Проверка атрибутивного состава
def check_the_structure_rdv(path):
    book = load_workbook(filename=path)
    shem_table = book['S2T таблицы']
    shem_field = book['S2T поля']
    values = [
        "'" + shem_table['E' + str(3)].value[13:29] + "'",
        "'" + shem_table['E' + str(4)].value[4:25] + "'",
        "'" + shem_table['F' + str(3)].value[4:25] + "'",
        "'" + shem_table['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    item = 3
    i = item
    hj = 0
    while ((shem_field['F' + str(item)].value) != None):
        if shem_field['C' + str(item)].value == 1:
            hj = hj + 1
        item = item + 1
    source_rdv = [о for о in range(hj)]
    source_type = [о for о in range(hj)]
    for j in range(i, i + hj):
        source_rdv[j - i] = shem_field['F' + str(j)].value
        source_type[j - i] = shem_field['I' + str(j)].value
    result = "with mapping_info as ("
    for j in range(i, i + hj):
        if j == i + hj - 1:
            result = result + "select '" + source_rdv[j - i] + "', '" + source_type[j - i] + "' as column_name\n"
        else:
            result = result + "select '" + source_rdv[j - i] + "', '" + source_type[
                j - i] + "' as column_name union all\n"
    result = result + ")select 'inf_schema', * from ( select column_name, udt_name from information_schema.columns where table_name = '" + \
             shem_table['E' + str(4)].value[4:25] + "'\n"
    result = result + "and table_schema = 'rdv'\n" \
                      "except\n" \
                      "select * from mapping_info ) t\n" \
                      "union all \n" \
                      "select 'mapping_info', * from (\n" \
                      "select * from mapping_info\n" \
                      "except\n" \
                      "select column_name, udt_name from information_schema.columns\n" \
                      "where table_name = " + values[1] + " and table_schema = 'rdv' ) t;"
    return result


# 4 Проверка логических дублей
def check_doubles(path):
    book = load_workbook(filename=path)
    shem_table = book['S2T таблицы']
    shem_field = book['S2T поля']
    item = 0
    i = 3
    while ((shem_field['C' + str(i)].value) == 1):
        if (shem_field['J' + str(i)].value == "PK") or (shem_field['J' + str(i)].value == "PK (History)"):
            item = item + 1
        i = i + 1
    values = [о for о in range(item)]
    item = 0
    i = 3
    while ((shem_field['C' + str(i)].value) == 1):
        if (shem_field['J' + str(i)].value == "PK") or (shem_field['J' + str(i)].value == "PK (History)"):
            values[item] = shem_field['F' + str(i)].value
            item = item + 1
        i = i + 1
    rezult = "select "
    for g in range(item):
        rezult = rezult + values[g]
        if g != item: rezult = rezult + ", "
    rezult = rezult + " count(*) from idl." + shem_table["F" + str(3)].value[4:25] + " group by "
    for g in range(item):
        rezult = rezult + values[g]
        if g != item - 1: rezult = rezult + ", "
    rezult = rezult + " having  count(*)>1;"
    return rezult


# 5 Проверка отсутствия старых объектов в базе
def absence_of_old(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        "'" + shem['E' + str(3)].value[13:29] + "'",
        "'" + shem['E' + str(4)].value[8:24] + "'",
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    results = "select count(*) from information_schema." + values[4] + " where \n" \
                                                                       "     lower(table_name) in (lower(map_" + values[
                  1] + "))\n" \
                       "     and table_type = 'TABLE' "
    return results


# 6 Проверка общего кол-ва строк эталона и мапа
def total_number_of_rows_reference_and_map(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        shem['E' + str(4)].value,
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        "'" + shem['D' + str(4)].value[8:30] + "'"
    ]
    results = "select count(*)- (\n" \
              " select count(*) from " + values[1] + ") \n" \
                                                     "     from rdv.mart_unified_map_acpd muma\n" \
                                                     "     where UPPER(map_table_name_dk) = UPPER(" + values[4] + ")\n" \
                                                                                                                  "     and map_schema_name_dk = '" + \
              shem['D' + str(3)].value[0:7] + "';"
    return results


# 7 Проверка истории (есть ли невалидные интервалы)
def check_history(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value[12:29],
        shem['E' + str(4)].value[8:25],
        "'" + shem['F' + str(3)].value[4:25] + "'",
        shem['G' + str(3)].value[4:25],
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    rdv = "select * from rdv.map_" + values[1] + " where effective_from_date = effective_to_date;"
    rdv_dict = "select * from rdv_dict.ref" + values[0] + " where effective_from_date = effective_to_date;"
    return rdv


# 8 Проверка нахождения дат в корректном интервале
def check_current_interval(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value[12:29],
        shem['E' + str(4)].value,
        "'" + shem['F' + str(3)].value[4:25] + "'",
        shem['G' + str(3)].value[4:25],
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    rezult = "select * from " + values[1] + " where 1 = 1\n" \
                                            "AND (effective_from_date < to_date('01.01.1900', 'dd.mm.yyyy') \n" \
                                            "OR effective_to_date > to_date('31.12.2999', 'dd.mm.yyyy')); "
    return rezult


# 9-10Проверка наличия аксессора типа "Вью" и типа "Inline-функция"
def availability_of_required_accessories_for_rdv1(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        shem['E' + str(4)].value[4:25],
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    result_view = "select count(*)=1 from information_schema.tables t where t.table_name in ('v_sn_vld_" + values[
        1] + "')\n" \
             "and t.table_schema = 'rdv' and t.table_type = 'VIEW';"

    result_func = "select count(*)=2 from information_schema.routines r where r.routine_name in ('v_sv_vld_" + values[
        1] + "', 'v_iv_vld_" + values[1] + "')\n" \
                                           "and r.routine_schema = 'rdv' and r.routine_type = 'FUNCTION';"
    return result_view


def availability_of_required_accessories_for_rdv2(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        shem['E' + str(4)].value[4:25],
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    result_view = "select count(*)=1 from information_schema.tables t where t.table_name in ('v_sn_vld_" + values[
        1] + "')\n" \
             "and t.table_schema = 'rdv' and t.table_type = 'VIEW';"

    result_func = "select count(*)=2 from information_schema.routines r where r.routine_name in ('v_sv_vld_" + values[
        1] + "', 'v_iv_vld_" + values[1] + "')\n" \
                                           "and r.routine_schema = 'rdv' and r.routine_type = 'FUNCTION';"
    return result_func


# 11-13 Запуск аксессоров
def check_pusk_accesor(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        shem['E' + str(4)].value[8:25],
        "'" + shem['F' + str(3)].value[4:25] + "'",
        shem['G' + str(3)].value[4:25],
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    r_sn_map = "select 1 from rdv.v_sn_vld_map_" + values[1] + " where 1=0;"
    r_sv_map = "select 1 from rdv.v_sv_vld_map_" + values[1] + " where 1=0;"
    r_iv_map = "select 1 from rdv.v_iv_vld_map_" + values[1] + " where 1=0;"
    v_sn_all_dict = "select 1 from rdv.v_sn_all_" + values[3] + " where 1=0;"
    v_r_dict = "select 1 from rdv.v_r_" + values[3] + " where 1=0;"
    dict = "select 1 from rdv." + values[3] + " where 1=0;"
    v_sv_all_dict = "select 1 from rdv.v_sv_all_" + values[3] + " where 1=0;"
    v_sv_hash_dict = "select 1 from rdv.v_sv_hash_" + values[3] + " where 1=0;"
    v_iv_all_dict = "select 1 from rdv.v_iv_all_" + values[3] + " where 1=0;"
    return r_sn_map, r_sv_map, r_iv_map, v_sn_all_dict, v_r_dict, dict, v_sv_all_dict, v_sv_hash_dict, v_iv_all_dict


def array_sripts_rdv(path):
    rez = [check_true_schems_rdv(path), check_data_in_directories_rdv(path), check_the_structure_rdv(path),
           check_doubles(path), absence_of_old(path), total_number_of_rows_reference_and_map(path),
           check_history(path), check_current_interval(path), availability_of_required_accessories_for_rdv1(path),
           availability_of_required_accessories_for_rdv2(path),
           check_pusk_accesor(path)]
    return rez


def get_array_objects(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(4)].value,
        shem['E' + str(3)].value,
        shem['F' + str(3)].value,
        shem['G' + str(3)].value,
    ]
    return values


# Проверки для слоя RDV_DICT

# 1 Проверка доступности таблицы
def check_true_schems_rdv_dict(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        "'" + shem['E' + str(3)].value[9:29] + "'",
        "'" + shem['E' + str(4)].value[4:24] + "'",
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    result = "select table_name from information_schema.tables where 1=1 \n" \
             "and table_schema = 'rdv_dict' \n" \
             "and table_name = " + values[0] + "\n" \
                                               "and table_type ='VIEW'"
    return result


# 2  Проверка наличия данных в справочниках
def check_data_in_directories_rdv_dict(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        "rdv.map_" + shem['E' + str(4)].value[8:24] + "",
        "idl." + shem['F' + str(3)].value[4:25] + "",
        "bdm." + shem['G' + str(3)].value[4:25] + "",
    ]
    results_rdv = "select count(*)>1 from " + values[1] + " where src_cd = 'ACPD' and src_cd" + shem[
        'D' + str(4)].value + ";"
    results_rdv_dict = "select * from " + values[0] + " where src_cd = 'ACPD' and src_cd='" + shem[
        'D' + str(3)].value + "';"
    results_idl = "select * from " + values[2] + " where src_cd = 'ACPD' and src_cd" + shem['D' + str(3)].value + ";"
    results_bdm = "select * from " + values[3] + " where src_cd = 'ACPD' and src_cd=" + shem['D' + str(3)].value + ";"
    return results_rdv_dict


# 3 Проверка атрибутивного состава
def check_the_structure_rdv_dict(path):
    book = load_workbook(filename=path)
    shem_table = book['S2T таблицы']
    shem_field = book['S2T поля']
    values = [
        "'" + shem_table['E' + str(3)].value[13:29] + "'",
        "'" + shem_table['E' + str(4)].value[4:25] + "'",
        "'" + shem_table['F' + str(3)].value[4:25] + "'",
        "'" + shem_table['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    item = 3
    i = item
    hj = 0
    while ((shem_field['F' + str(item)].value) != None):
        if shem_field['C' + str(item)].value == 2:
            hj = hj + 1
        item = item + 1
    source_rdv = [о for о in range(hj)]
    source_type = [о for о in range(hj)]
    for j in range(item - hj, item):
        source_rdv[j - hj] = shem_field['F' + str(j)].value
        source_type[j - hj] = shem_field['I' + str(j)].value
    result = "with mapping_info as ("
    for j in range(i, i + hj):
        if j == i + hj - 1:
            result = result + "select '" + source_rdv[j - i] + "', '" + source_type[j - i] + "' as column_name\n"
        else:
            result = result + "select '" + source_rdv[j - i] + "', '" + source_type[
                j - i] + "' as column_name union all\n"
    result = result + ")select 'inf_schema', * from ( select column_name, udt_name from information_schema.columns where table_name = '" + \
             shem_table['E' + str(3)].value[9:29] + "'\n"
    result = result + "and table_schema = 'rdv_dict'\n" \
                      "except\n" \
                      "select * from mapping_info ) t\n" \
                      "union all \n" \
                      "select 'mapping_info', * from (\n" \
                      "select * from mapping_info\n" \
                      "except\n" \
                      "select column_name, udt_name from information_schema.columns\n" \
                      "where table_name = " + values[0] + " and table_schema = 'rdv_dict' ) t;"
    return result


# 4Проверка логических дублей
def check_doubles_rdv_dict(path):
    book = load_workbook(filename=path)
    shem_table = book['S2T таблицы']
    shem_field = book['S2T поля']
    item = 3
    i = 3
    hj = 0
    while ((shem_field['F' + str(item)].value) != None):
        if (shem_field['C' + str(item)].value == 2) and (
                (shem_field['J' + str(item)].value == "PK") or (shem_field['J' + str(item)].value == "PK (History)")):
            hj = hj + 1
        item = item + 1
    values = [о for о in range(hj)]
    item = 0
    i = 3
    while ((shem_field['F' + str(i)].value) != None):
        if (shem_field['C' + str(i)].value == 2) and (
                (shem_field['J' + str(i)].value == "PK") or (shem_field['J' + str(i)].value == "PK (History)")):
            values[item] = shem_field['F' + str(i)].value
            item = item + 1
        i = i + 1

    rezult = "select "
    for g in range(item):
        rezult = rezult + values[g]
        if g != item: rezult = rezult + ", "
    rezult = rezult + " count(*) from rdv_dict.ref_" + shem_table["E" + str(3)].value[13:29] + " group by "
    for g in range(item):
        rezult = rezult + values[g]
        if g != item - 1: rezult = rezult + ", "
    rezult = rezult + " having  count(*)>1;"
    return rezult


# 5Проверка истории (есть ли невалидные интервалы)
def check_history_rdv_dict(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value[12:29],
        shem['E' + str(4)].value[8:25],
        "'" + shem['F' + str(3)].value[4:25] + "'",
        shem['G' + str(3)].value[4:25],
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    rdv = "select * from rdv.map_" + values[1] + " where effective_from_date = effective_to_date;"
    rdv_dict = "select * from rdv_dict.ref" + values[0] + " where effective_from_date = effective_to_date;"
    return rdv_dict


# 6Проверка нахождения дат в корректном интервале
def check_current_interval_rdv_dict(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        shem['E' + str(4)].value,
        "'" + shem['F' + str(3)].value[4:25] + "'",
        shem['G' + str(3)].value[4:25],
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    rezult = "select * from " + values[0] + " where 1 = 1\n" \
                                            " AND (effective_from_date < to_date('01.01.1900', 'dd.mm.yyyy') \n" \
                                            "OR effective_to_date > to_date('31.12.2999', 'dd.mm.yyyy')); "
    return rezult


# 7Проверка отсутствия старых объектов в базе
def absence_of_old_rdv_dict(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        "'" + shem['E' + str(3)].value[9:29] + "'",
        "'" + shem['E' + str(4)].value[8:24] + "'",
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    results = "select count(*) = 0 from information_schema." + values[4] + " where \n" \
                                                                           "     lower(table_name) in (lower(" + values[
                  0] + "))\n" \
                       "     and table_type = 'TABLE' "
    return results


# 8Проверка общего кол-ва строк эталона и
def total_number_of_rows_reference_and_map_rdv_dict(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        shem['E' + str(4)].value,
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    results = "select count(*)- (\n" \
              " select count(*) from " + values[0] + ") \n" \
                                                     "     from rdv.mart_unified_ref_acpd muma\n" \
                                                     "     where UPPER(ref_table_name_dk) = UPPER(" + values[4] + ")\n" \
                                                                                                                  "     and ref_schema_name_dk = '" + \
              shem['D' + str(3)].value[0:7] + "';"
    return results


def array_sripts_rdv_dict(path):
    rez = [check_true_schems_rdv_dict(path), check_data_in_directories_rdv_dict(path),
           check_the_structure_rdv_dict(path),
           check_doubles_rdv_dict(path), absence_of_old_rdv_dict(path),
           total_number_of_rows_reference_and_map_rdv_dict(path),
           check_history_rdv_dict(path), check_current_interval_rdv_dict(path)]
    return rez


# Проверки для слоя IDL
# 1Проверка доступности таблицы
def check_true_schems_IDL(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        "'" + shem['E' + str(3)].value[9:29] + "'",
        "'" + shem['E' + str(4)].value[4:24] + "'",
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    result = "select table_name from information_schema.tables where 1=1 \n" \
             "and table_schema = 'idl' \n" \
             "and table_name = " + values[2] + "\n" \
                                               "and table_type ='VIEW'"
    return result


# 2	Проверка наличия данных в справочниках
def check_data_in_directories_IDL(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        "rdv.map_" + shem['E' + str(4)].value[8:24] + "",
        "idl." + shem['F' + str(3)].value[4:25] + "",
        "bdm." + shem['G' + str(3)].value[4:25] + "",
    ]
    results_rdv = "select count(*)>1 from " + values[1] + " where src_cd = 'ACPD' and src_system=" + shem[
        'D' + str(4)].value + ";"
    results_rdv_dict = "select * from " + values[0] + " where src_cd = 'ACPD' and src_system='" + shem[
        'D' + str(3)].value + "';"
    results_idl = "select * from " + values[2] + " where src_cd = 'ACPD' and src_cd='" + shem['D' + str(3)].value + "';"
    results_bdm = "select * from " + values[3] + " where src_cd = 'ACPD' and src_system=" + shem[
        'D' + str(3)].value + ";"
    return results_idl


# 3Проверка атрибутивного состава
def check_the_structure_IDL(path):
    book = load_workbook(filename=path)
    shem_table = book['S2T таблицы']
    shem_field = book['S2T поля']
    values = [
        "'" + shem_table['E' + str(3)].value[13:29] + "'",
        "'" + shem_table['E' + str(4)].value[4:25] + "'",
        "'" + shem_table['F' + str(3)].value[4:25] + "'",
        "'" + shem_table['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    item = 3
    i = item
    hj = 0
    if shem_field['G' + str(3)].value == None: item = item + 1
    i = item
    while ((shem_field['G' + str(item)].value) != None):
        if shem_field['C' + str(item)].value == 1:
            hj = hj + 1
        item = item + 1
    source_idl = [о for о in range(hj)]
    source_type = [о for о in range(hj)]
    for j in range(i, i + hj):
        source_idl[j - i] = shem_field['G' + str(j)].value
        source_type[j - i] = shem_field['I' + str(j)].value
    result = "with mapping_info as ("
    for j in range(i, i + hj):
        if j == i + hj - 1:
            result = result + "select '" + source_idl[j - i] + "', '" + source_type[j - i] + "' as column_name\n"
        else:
            result = result + "select '" + source_idl[j - i] + "', '" + source_type[
                j - i] + "' as column_name union all\n"
    result = result + ")select 'inf_schema', * from ( select column_name, udt_name from information_schema.columns where table_name = '" + \
             shem_table['E' + str(4)].value[4:25] + "'\n"
    result = result + "and table_schema = 'rdv'\n" \
                      "except\n" \
                      "select * from mapping_info ) t\n" \
                      "union all \n" \
                      "select 'mapping_info', * from (\n" \
                      "select * from mapping_info\n" \
                      "except\n" \
                      "select column_name, udt_name from information_schema.columns\n" \
                      "where table_name = " + values[2] + " and table_schema = 'idl' ) t;"
    return result


# 4Заполнение src_cd
def filling_in_src_cd(path):
    book = load_workbook(filename=path)
    shem_table = book['S2T таблицы']
    shem_field = book['S2T поля']
    values = [
        shem_table['E' + str(3)].value,
        "'" + shem_table['E' + str(4)].value[4:25] + "'",
        shem_table['F' + str(3)].value,
        "'" + shem_table['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    result = "select * from " + values[2] + " mrt\n" \
                                            "where not exists (select 1 from " + values[
                 0] + " ss where mrt.src_cd = ss.src_cd);"
    return result


# 5Проверка пересечений и разрывов в истории:
def check_intersections_and_gaps_in_history(path):
    book = load_workbook(filename=path)
    shem_table = book['S2T таблицы']
    shem_field = book['S2T поля']
    values = [
        shem_table['E' + str(3)].value,
        "'" + shem_table['E' + str(4)].value[4:25] + "'",
        shem_table['F' + str(3)].value[9:25],
        "'" + shem_table['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    rezult = "with acc_res as ( select * from( select * ,\n" \
             "row_number() over (partition by " + values[
                 2] + "_cd, effective_from_date order by version_id desc) as rn\n" \
                      "from idl.v_sn_all_dict_" + values[2] + ") as a where rn = 1) select * from (\n" \
                                                              "select " + values[
                 2] + "_cd, effective_from_date, effective_to_date,\n" \
                      "lead(effective_from_date) over (partition by " + values[
                 2] + "_cd order by effective_from_date, effective_to_date) as lead_eff_from_dt\n" \
                      "from  acc_res order by " + values[2] + "_cd, effective_from_date\n" \
                                                              ") res where effective_to_date != lead_eff_from_dt;"
    return rezult


# 6Проверка отсутствия старых объектов в базе
def absence_of_old_IDL(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        "'" + shem['E' + str(3)].value[9:29] + "'",
        "'" + shem['E' + str(4)].value[8:24] + "'",
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    results = "select count(*) = 0 from information_schema." + values[4] + " where \n" \
                                                                           "     lower(table_name) in (lower(" + values[
                  2] + "))\n" \
                       "     and table_type = 'TABLE' "
    return results


def array_sripts_IDL(path):
    rez = [check_true_schems_IDL(path), check_data_in_directories_IDL(path), check_the_structure_IDL(path),
           filling_in_src_cd(path), check_intersections_and_gaps_in_history(path), absence_of_old_IDL(path)]
    return rez


# Проверки для слоя BDM

# 1Проверка доступности таблицы
def check_true_schems_BDM(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        "'" + shem['E' + str(3)].value[9:29] + "'",
        "'" + shem['E' + str(4)].value[4:24] + "'",
        "'" + shem['F' + str(3)].value[4:25] + "'",
        "'" + shem['G' + str(3)].value[4:25] + "'",
        '"tables"'
    ]
    result = "select table_name from information_schema.tables where 1=1 \n" \
             "and table_schema = 'BDM' \n" \
             "and table_name = " + values[2] + "\n" \
                                               "and table_type ='VIEW'"
    return result


# 2Проверка наличия данных в справочниках
def check_data_in_directories_BDM(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        "rdv.map_" + shem['E' + str(4)].value[8:24] + "",
        "idl." + shem['F' + str(3)].value[4:25] + "",
        "bdm." + shem['G' + str(3)].value[4:25] + "",
    ]
    results_rdv = "select count(*)>1 from " + values[1] + " where src_cd = 'ACPD' and src_system=" + shem[
        'D' + str(4)].value + ";"
    results_rdv_dict = "select * from " + values[0] + " where src_cd = 'ACPD' and src_system='" + shem[
        'D' + str(3)].value + "';"
    results_idl = "select * from " + values[2] + " where src_cd = 'ACPD' and src_system=" + shem[
        'D' + str(3)].value + ";"
    results_bdm = "select * from " + values[3] + " where src_cd = 'ACPD' and src_cd='" + shem['D' + str(3)].value + "';"
    return results_bdm


# 3-4 Проверка наличия аксессоров типа "Вью" и типа "Inline-функция"
def availability_of_required_accessories_for_bdm(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        shem['E' + str(4)].value[4:25],
        "'" + shem['F' + str(3)].value[4:25] + "'",
        shem['G' + str(3)].value[4:25],
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    result_view = "select count(*)=3 from information_schema.tables t\n" \
                  "where t.table_name in ('v_sn_all_" + values[3] + "', 'v_r_" + values[3] + "', '" + values[3] + "')\n" \
                                                                                                                  "and t.table_schema = 'bdm' and t.table_type = 'VIEW';\n"
    result_func = "select count(*)=3 from information_schema.routines r where r.routine_name in ('v_sv_all_" + values[
        3] + "', 'v_sv_hash_" + values[3] + "', 'v_iv_all_" + values[3] + "')\n" \
                                                                          "and r.routine_schema = 'bdm' and r.routine_type = 'FUNCTION';"
    return result_view, result_func


# 5Запуск аксессоров
def check_pusk_accesor_BDM(path):
    book = load_workbook(filename=path)
    shem = book['S2T таблицы']
    values = [
        shem['E' + str(3)].value,
        shem['E' + str(4)].value[8:25],
        "'" + shem['F' + str(3)].value[4:25] + "'",
        shem['G' + str(3)].value[4:25],
        "'" + shem['D' + str(3)].value[8:30] + "'"
    ]
    v_sn_all_dict = "select 1 from bdm.v_sn_all_" + values[3] + " where 1=0;"
    v_r_dict = "select 1 from bdm.v_r_" + values[3] + " where 1=0;"
    dict = "select 1 from bdm." + values[3] + " where 1=0;"
    v_sv_all_dict = "select 1 from bdm.v_sv_all_" + values[3] + " where 1=0;"
    v_sv_hash_dict = "select 1 from bdm.v_sv_hash_" + values[3] + " where 1=0;"
    v_iv_all_dict = "select 1 from bdm.v_iv_all_" + values[3] + " where 1=0;"
    return v_sn_all_dict, v_r_dict, dict, v_sv_all_dict, v_sv_hash_dict, v_iv_all_dict


def array_sripts_BDM(path):
    rez = [check_true_schems_BDM(path), check_data_in_directories_BDM(path),
           availability_of_required_accessories_for_bdm(path)[0],
           availability_of_required_accessories_for_bdm(path)[1], check_pusk_accesor_BDM(path)[0],
           check_pusk_accesor_BDM(path)[1],
           check_pusk_accesor_BDM(path)[2], check_pusk_accesor_BDM(path)[3],
           check_pusk_accesor_BDM(path)[4], check_pusk_accesor_BDM(path)[5]]
    return rez
